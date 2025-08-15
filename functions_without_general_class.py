import io
import os
import shutil
from typing import Any
import json
import openpyxl

import variables
from class_file import StatusFile, ClassFile


def get_list_of_all_protocols(dir_protocols: str) -> tuple[list[Any] | list[str], list[Any]]:
    all_files = get_only_files(path=dir_protocols)
    doc = variables.types_of_the_protocol_files[0]
    docx = variables.types_of_the_protocol_files[1]
    doc_files = {x[:-len(doc)] for x in all_files if x[-len(doc):] == doc}
    docx_files = {x[:-len(docx)] for x in all_files if x[-len(docx):] == docx}
    filter_files = doc_files | docx_files
    filter_files_2 = set()
    for variant_of_protocol in variables.names_of_protocol:
        for file_name_i in filter_files:
            if variant_of_protocol in file_name_i:
                n = file_name_i.index(variant_of_protocol)
                file_filter_name = file_name_i[n + len(variant_of_protocol):]
                filter_files_2.add(file_filter_name)
    send_protocols = []
    not_send_protocols = []
    for protokol_i in filter_files_2:
        if protokol_i.isnumeric():
            send_protocols.append(variables.protocol + '_' + protokol_i)
        else:
            not_send_protocols.append(variables.protocol + protokol_i)
    if len(not_send_protocols) == 0:
        not_send_protocols = [variables.protocol + '_' + '00']
    return not_send_protocols, send_protocols


def start_file_is_send(path: str, name: str, subdir: str):
    path = path.replace('\\' + name, '')
    if subdir != '':
        path = path.replace('\\' + subdir, '')
    path = path.replace(variables.incoming_docs, variables.checked_files)
    all_files = get_only_files(path=path)
    for ending in variables.variants_of_the_ending:
        name_i = name.replace('.pdf', '_' + ending + '.pdf')
        if name_i in all_files:
            path += '\\' + name_i
            start_the_file(path=path)


def check_the_file(name: str, dict_i: dict, folder: str, status_name: str) -> tuple[str, int] | bool:
    for protocol_nr_i, dict_of_files_i in dict_i.items():
        if folder == '':
            if protocol_nr_i == 0:
                if name in dict_i[protocol_nr_i]:
                    return status_name, protocol_nr_i
            else:
                if name in dict_i[protocol_nr_i][0]:
                    return status_name, protocol_nr_i
            continue
        else:
            if isinstance(dict_of_files_i, list):
                continue
            for folder_j, list_of_files_j in dict_of_files_i.items():
                if folder_j == folder:
                    if name in list_of_files_j:
                        return status_name, protocol_nr_i
    return False


def get_dict_to_send_files(path: str, dict_checked_files: dict) -> dict:
    if len(dict_checked_files) == 0:
        return dict()
    path = path + '\\' + variables.checked_files
    all_folder = get_only_folders(path=path)
    if variables.files_to_send not in set(all_folder):
        print('no subdir', variables.files_to_send)
        return dict()
    path = path + '\\' + variables.files_to_send
    all_folder = get_only_folders(path=path)
    return get_dict_with_protocol_files(path=path, folders=all_folder)


def get_dict_of_checked_files(path: str, dict_by_checking: dict) -> dict:
    if len(dict_by_checking) == 0:
        return dict()
    path = path + '\\' + variables.checked_files
    all_folder = get_only_folders(path=path)
    if variables.checked_files_planes not in set(all_folder):
        print('no subdir', variables.checked_files_planes)
        return dict()
    path = path + '\\' + variables.checked_files_planes
    all_folder = get_only_folders(path=path)
    return get_dict_with_protocol_files(path=path, folders=all_folder)


def get_dict_of_by_checking_files(path: str) -> dict:
    all_folder = get_only_folders(path=path)
    if variables.checked_files not in set(all_folder):
        print('no subdir !!', variables.checked_files)
        return dict()
    path = path + '\\' + variables.checked_files
    all_folder = get_only_folders(path=path)
    if variables.by_checking not in set(all_folder):
        print('no subdir', variables.by_checking)
        return dict()
    path = path + '\\' + variables.by_checking
    all_folder = get_only_folders(path=path)
    return get_dict_with_protocol_files(path=path, folders=all_folder)


def get_dict_with_protocol_files(path: str, folders: list[str]) -> dict:
    # all files in folders
    dict_protokol_files = dict()

    for folder_i in folders:
        for name_i in variables.names_of_protocol:
            if folder_i.find(name_i) != -1:
                nummer_i = folder_i.replace(name_i, '')
                try:
                    nummer_i = int(nummer_i)
                except Exception as err:
                    print('there is no number', folder_i)
                    print(f"Unexpected {err=}, {type(err)=}")
                    raise
                path_i = path + '\\' + folder_i
                dict_protokol_files[nummer_i] = {0:get_only_files(path=path_i)}
                folders_ii = get_only_folders(path=path_i)
                for folder_j in folders_ii:
                    path_j = path_i + '\\' + folder_j
                    dict_protokol_files[nummer_i][folder_j] = get_only_files(path=path_j)
                continue
    return dict_protokol_files


def get_list_of_send_files(path: str) -> list[str]:
    all_folder = get_only_folders(path=path)
    if variables.checked_files not in set(all_folder):
        print('no subdir', variables.checked_files)
        return []
    path = path + '\\' + variables.checked_files
    row_list = set(get_only_files(path=path))
    list_of_folders = set(get_only_folders(path=path))

    list_of_folders = list_of_folders.difference(variables.file_name_not_to_scan)
    for folder in list_of_folders:
        path_i = path + '\\' + folder
        files_in_the_folder = set(get_only_files(path=path_i))
        row_list = row_list.union(files_in_the_folder)
    list_of_files = []
    for file in row_list:
        if file[-4:] == '.pdf':
            file = file[:-4]
            for n in [7, 6, 4]:
                if file[-n:] in variables.variants_of_the_ending:
                    list_of_files.append(file[:-(n + 1)])
                    break
    return list_of_files


def get_only_folders(path: str) -> list[str]:
    """print("Only directories:")
    print([name for name in os.listdir(path) if os.path.isdir(os.path.join(path, name))])
    print("\nOnly files:")
    print([name for name in os.listdir(path) if not os.path.isdir(os.path.join(path, name))])
    print("\nAll directories and files :")
    print([name for name in os.listdir(path)])"""
    return [name for name in os.listdir(path) if os.path.isdir(os.path.join(path, name))]


def get_only_files(path: str) -> list[str]:
    return [name for name in os.listdir(path) if not os.path.isdir(os.path.join(path, name))]


def start_file_by_status(path: str, name: str, protokol_nr: int, folder_to_check: str, subdir: str):
    path = path.replace(variables.incoming_docs, variables.checked_files + '\\' + folder_to_check)
    path = path.replace(name, '')
    if subdir != '':
        path = path.replace('\\' + subdir, '')
    all_folder = get_only_folders(path=path)
    for name_i in variables.names_of_protocol:
        name_of_the_protocol = name_i + str(protokol_nr)
        if name_of_the_protocol in all_folder:
            if path[-1:] == "\\":
                path = path[:-1]
            path += '\\' + name_of_the_protocol
            if subdir != '':
                path += '\\' + subdir
            all_files = get_only_files(path=path)
            if name in all_files:
                path += '\\' + name
                start_the_file(path=path)


def start_the_file(path: str):
    os.startfile(path)


def copy_the_file(old_path: str, new_path: str, name: str) -> None:
    print('I copy the file', name)
    print('from', old_path)
    print('to', new_path)
    print('---->>>>>')

    if os.path.isfile(old_path) is False:
        print('there is no file', name)
        return None
    new_path_for_the_folder = new_path.replace(name, '')
    if not os.path.exists(new_path_for_the_folder):
        os.makedirs(new_path_for_the_folder)
    try:
        shutil.copyfile(old_path, new_path)
        return None
    except Exception as err:
        print('copy error', name)
        print(f"Unexpected {err=}, {type(err)=}")
        raise


def move_the_file(old_path: str, new_path: str, name: str) -> None:
    print('I move the file', name)
    print('from', old_path)
    print('to', new_path)
    print('---->>>>>')

    if os.path.isfile(old_path) is False:
        print('there is no file', name)
        return None
    new_path_for_the_folder = new_path.replace(name, '')
    if not os.path.exists(new_path_for_the_folder):
        os.makedirs(new_path_for_the_folder)
    try:
        shutil.move(old_path, new_path)
        return None
    except Exception as err:
        print('move error', name)
        print(f"Unexpected {err=}, {type(err)=}")
        raise


def move_from_unchecked_to_by_checking(file: ClassFile, protocol: str = ''):
    old_path = file.path
    if protocol == '':
        new_part = variables.checked_files + '\\' + variables.by_checking
    else:
        new_part = variables.checked_files + '\\' + variables.by_checking + '\\' + protocol
    new_path = file.path.replace(variables.incoming_docs, new_part)
    file.status = StatusFile.by_checking
    copy_the_file(old_path=old_path, new_path=new_path, name=file.name)


def move_from_by_checking_to_checked(file: ClassFile, protocol: str = ''):
    if protocol == '':
        new_part = variables.checked_files + '\\' + variables.by_checking
    else:
        new_part = variables.checked_files + '\\' + variables.by_checking + '\\' + protocol
    old_path = file.path.replace(variables.incoming_docs, new_part)

    if protocol == '':
        new_part = variables.checked_files + '\\' + variables.checked_files_planes
    else:
        new_part = variables.checked_files + '\\' + variables.checked_files_planes + '\\' + protocol
    new_path = file.path.replace(variables.incoming_docs, new_part)
    copy_the_file(old_path=old_path, new_path=new_path, name=file.name)
    file.status = StatusFile.checked


def move_from_checked_to_to_send(file: ClassFile, protocol: str = ''):
    if protocol == '':
        new_part = variables.checked_files + '\\' + variables.checked_files_planes
    else:
        new_part = variables.checked_files + '\\' + variables.checked_files_planes + '\\' + protocol
    old_path = file.path.replace(variables.incoming_docs, new_part)
    if protocol == '':
        new_part = variables.checked_files + '\\' + variables.files_to_send
    else:
        new_part = variables.checked_files + '\\' + variables.files_to_send + '\\' + protocol
    new_path = file.path.replace(variables.incoming_docs, new_part)
    copy_the_file(old_path=old_path, new_path=new_path, name=file.name)
    file.status = StatusFile.to_send


def move_the_file_to_new_protocol(file: ClassFile, new_protocol: str):
    match file.status:
        case StatusFile.by_checking:
            list_of_status = [StatusFile.by_checking]
        case StatusFile.checked:
            list_of_status = [StatusFile.by_checking, StatusFile.checked]
        case StatusFile.to_send | StatusFile.is_send:
            list_of_status = [StatusFile.by_checking, StatusFile.checked, StatusFile.to_send]
        case _:
            list_of_status = []
    for status in list_of_status:
        move_the_file_to_new_protocol_with_status(file=file, new_protocol=new_protocol,
                                                  status=status)


def move_the_file_to_new_protocol_with_status(file: ClassFile, new_protocol: str, status: str):
    old_protocol = variables.protocol + ' ' + str(file.nr_protokol)
    match status:
        case StatusFile.by_checking:
            status_path = variables.by_checking
        case StatusFile.checked:
            status_path = variables.checked_files_planes
        case StatusFile.to_send:
            status_path = variables.files_to_send
        case _:
            status_path = variables.by_checking
    move_the_file_with_the_status(file=file, new_protocol=new_protocol, old_protocol=old_protocol,
                                  status_path=status_path)


def move_the_file_with_the_status(file: ClassFile, new_protocol: str, status_path: str, old_protocol: str):
    new_part = variables.checked_files + '\\' + status_path + '\\' + old_protocol
    old_path = file.path.replace(variables.incoming_docs, new_part)
    new_part = variables.checked_files + '\\' + status_path + '\\' + str(new_protocol)
    new_path = file.path.replace(variables.incoming_docs, new_part)
    move_the_file(old_path=old_path, new_path=new_path, name=file.name)


def open_the_file(file: ClassFile):
    match file.status:
        case StatusFile.unchecked:
            start_the_file(path=file.path)
        case StatusFile.by_checking:
            start_file_by_status(path=file.path, name=file.name, protokol_nr=file.nr_protokol,
                                 folder_to_check=variables.by_checking, subdir=file.subdir)
        case StatusFile.checked:
            start_file_by_status(path=file.path, name=file.name, protokol_nr=file.nr_protokol,
                                 folder_to_check=variables.checked_files_planes, subdir=file.subdir)
        case StatusFile.to_send:
            start_file_by_status(path=file.path, name=file.name, protokol_nr=file.nr_protokol,
                                 folder_to_check=variables.files_to_send, subdir=file.subdir)
        case StatusFile.is_send:
            start_file_is_send(path=file.path, name=file.name, subdir=file.subdir)


def get_list_of_file_in_the_protocol(list_of_all_files: list[ClassFile], number_of_current_protocol: int,
                                     current_dir_incoming_docs: str) ->  list[ClassFile]|None:
    list_of_file_in_protocol = []
    for file in list_of_all_files:
        if file.nr_protokol == number_of_current_protocol:
            list_of_file_in_protocol.append(file)
    all_files = get_only_files(path=current_dir_incoming_docs)
    all_excel_files = [name for name in all_files if name[-5:] == '.xlsx']
    file_name = ''
    for file in all_excel_files:
        if file.find(variables.text_of_the_excel_file):
            file_name = file
            break
    if file_name == '':
        return None
    dir_for_excel_file = current_dir_incoming_docs + '\\' + file_name
    dict_files = file_excel_open_get_date(file_name=dir_for_excel_file)
    if dict_files is None:
        return None
    return get_new_list_of_files_in_the_protocol(dict_files=dict_files, list_of_files=list_of_file_in_protocol)


def get_new_list_of_files_in_the_protocol(dict_files: dict[str:str], list_of_files: list[ClassFile]) -> list[ClassFile]:
    new_list_of_file = []
    dict_of_names = dict()
    for file_table in dict_files:
        if len(file_table) < 2:
            continue
        try:
            index = file_table.index(variables.text_index)
            file_name = file_table[:index].replace(" ", "")
            index = file_table[index + 7:].replace(" ", "")
            if index == '-':
                index = ''
        except ValueError:
            file_name = file_table
            index = ''
        value = dict_files[file_table]
        if file_name in dict_of_names:  # check the new index
            if index > dict_of_names[file_name][1]:
                dict_of_names[file_name] = [index, value]
            else:
                continue
        else:  # there is no file in the dict
            dict_of_names[file_name] = [index, value]
    for file in list_of_files:
        for file_name, values in dict_of_names.items():
            index, value = values

            if file.name.find(file_name) > 0:
                file_2 = ClassFile(name=file.name, status=file.status, nr_protokol=file.nr_protokol, subdir=file.subdir,
                                   path=file.path)
                file_2.name_of_file_in_the_table = file_name
                file_2.name_of_the_plan = value
                file_2.index = index
                new_list_of_file.append(file_2)
    return new_list_of_file


def file_excel_open_get_date(file_name: str) -> dict[str:str] | None:
    with (open(file_name, "rb") as f):
        try:
            in_mem_file = io.BytesIO(f.read())
        except FileExistsError:
            print('can not open the file')
            return None
        print("open the excel file ", file_name)
        work_book = openpyxl.load_workbook(in_mem_file, read_only=True)
        print("read the work_book ", work_book)
        print("with the sheetnames ", work_book.sheetnames)
        sheet = work_book.active

        column = 1
        while sheet.cell(row=1, column=column).value is not None:
            value = sheet.cell(row=1, column=column).value
            if value == variables.name_of_the_useful_cell_in_the_excel_file:
                break
            column+=1

        i = 2
        dict_files = dict()
        while sheet.cell(row=i, column=1).value is not None:
            name_index = str(sheet.cell(row=i, column=2).value)
            value = str(sheet.cell(row=i, column=column).value)
            value = '' if value == 'None' else value
            i += 1
            dict_files[name_index] = value
        work_book.close()
    return dict_files


def read_the_setting_from_the_file(path: str) -> dict[str:str]|None:
    try:
        # Open and read the file
        with open(path, 'r') as file:
            content = file.read()
        # Try to parse the content as JSON
        try:
            data_dict = json.loads(content)
            print("JSON data successfully parsed as dictionary:")
            print(data_dict)
            return {'dir_for_checking': data_dict['dir_for_checking'],
                    'year': data_dict['year'],
                    'project': data_dict['project'],
                    'show_static': data_dict['show_static']}
        except json.JSONDecodeError:
            print("The file content is not valid JSON.")
            return None
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None

def make_default_settings(path: str) -> dict[str:str]|None:
    print(f"The file does not exist at: {path}")
    dir_0 = variables.dir_for_checking
    raw_list_of_all_folder = os.listdir(dir_0)
    list_of_years = list(filter(lambda x: (x[:4] == variables.name_of_the_folder), raw_list_of_all_folder))
    current_year = list_of_years[-1]
    current_list_of_files_for_the_year: list[str] = get_only_folders(path=dir_0 + '\\' + current_year)
    project = current_list_of_files_for_the_year[0]
    settings_0 = {'dir_for_checking': variables.dir_for_checking,
                  'year': current_year,
                  'project': project,
                  'show_static': False}
    try:
        with open(path, 'w') as file:
            json.dump(settings_0, file, indent=4)
        print(f"Dictionary saved as JSON to '{path}'.")
    except Exception as e:
        print(f"Failed to save JSON: {e}")
    return settings_0